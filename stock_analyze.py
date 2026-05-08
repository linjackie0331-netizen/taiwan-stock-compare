#!/usr/bin/env python3
"""
taiwan-stock-compare: 台灣股票多股財務比較分析工具
用法：
  python stock_analyze.py 2317              # 單股詳細分析
  python stock_analyze.py 2317 2330 2454    # 多股同業比較
  python stock_analyze.py 2317 --years 5   # 指定年數
  python stock_analyze.py 2317 --no-cache  # 強制重新抓取
"""

import sys
import time
import sqlite3
import argparse
from datetime import date
from pathlib import Path

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

BASE_DIR   = Path(__file__).parent
CACHE_DB   = BASE_DIR / "cache.db"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

STOCK_COLORS = ['#3182ce', '#38a169', '#dd6b20', '#805ad5', '#e53e3e']

# ─── 快取層 ────────────────────────────────────────────────

def init_cache():
    conn = sqlite3.connect(CACHE_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS stock_cache (
            stock_id TEXT, rpt_cat TEXT, fetch_date TEXT, html TEXT,
            PRIMARY KEY (stock_id, rpt_cat, fetch_date)
        )
    """)
    conn.commit()
    return conn

# ─── 抓取層 ────────────────────────────────────────────────

def get_client_key():
    tz_offset = -480
    now_ms = time.time() * 1000
    days = now_ms / 86400000 - tz_offset / 1440
    return f"2.8|38057.1435627105|46946.0324515993|{tz_offset}|{days}|{days}", days

def _fetch_from_goodinfo(stock_id, rpt_cat, days, client_key):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://goodinfo.tw/'
    }
    url = (f"https://goodinfo.tw/tw/StockFinDetail.asp"
           f"?RPT_CAT={rpt_cat}&STOCK_ID={stock_id}&REINIT={days:.10f}")
    r = requests.get(url, headers=headers, cookies={'CLIENT_KEY': client_key}, timeout=20)
    r.encoding = 'utf-8'
    return r.text

def fetch_report(stock_id, rpt_cat, days, client_key, conn, use_cache=True):
    today = date.today().isoformat()
    if use_cache:
        row = conn.execute(
            "SELECT html FROM stock_cache WHERE stock_id=? AND rpt_cat=? AND fetch_date=?",
            (stock_id, rpt_cat, today)
        ).fetchone()
        if row:
            return BeautifulSoup(row[0], 'html.parser')
    html = _fetch_from_goodinfo(stock_id, rpt_cat, days, client_key)
    conn.execute("INSERT OR REPLACE INTO stock_cache VALUES (?,?,?,?)",
                 (stock_id, rpt_cat, today, html))
    conn.commit()
    return BeautifulSoup(html, 'html.parser')

# ─── 解析層 ────────────────────────────────────────────────
# 修正：舊版用 idx = j*2 假設每年固定佔 2 欄，實際上 Goodinfo 欄位不固定。
# 新版先掃第一行找出每個年份的精確欄位 index，再用那個 index 取值。

def parse_table(soup, max_years=5):
    tables = soup.find_all('table')
    if len(tables) < 7:
        return {}, []
    rows = tables[6].find_all('tr')
    years, year_cols, data = [], {}, {}

    for row in rows:
        cells = row.find_all(['td', 'th'])
        if not cells:
            continue
        row_data = [c.get_text(strip=True) for c in cells]

        if not years:
            col_map, found = {}, []
            for idx, val in enumerate(row_data):
                if len(val) == 4 and val.isdigit() and 2015 <= int(val) <= 2030:
                    if val not in col_map:
                        col_map[val] = idx
                        found.append(val)
                        if len(found) >= max_years:
                            break
            if found:
                years, year_cols = found, col_map
            continue

        if years and len(row_data) >= 2 and row_data[0]:
            values = {}
            for yr, col_idx in year_cols.items():
                if col_idx < len(row_data):
                    try:
                        values[yr] = float(row_data[col_idx].replace(',', ''))
                    except Exception:
                        values[yr] = None
            if any(v is not None for v in values.values()):
                data[row_data[0]] = values

    return data, years

# ─── 指標計算層 ────────────────────────────────────────────

def _g(table, key, yr):
    return table.get(key, {}).get(yr)

def _find(table, *keywords):
    for key in table:
        if all(kw in key for kw in keywords):
            return key
    return None

def safe_div(a, b, pct=True):
    if a is not None and b and b != 0:
        return a / b * (100 if pct else 1)
    return None

def calc_cagr(v_start, v_end, n):
    if v_start and v_end and v_start > 0 and n > 0:
        return ((v_end / v_start) ** (1 / n) - 1) * 100
    return None

def calculate_metrics(is_d, bs_d, cf_d, years):
    metrics = {}
    for yr in years:
        rev_k   = _find(is_d, '營業收入')
        gp_k    = _find(is_d, '毛利') or _find(is_d, '營業毛利')
        op_k    = _find(is_d, '營業利益')
        ni_k    = _find(is_d, '稅後淨利') or _find(is_d, '本期淨利')
        eps_k   = _find(is_d, '每股') or _find(is_d, 'EPS')
        ca_k    = _find(bs_d, '流動資產合計') or _find(bs_d, '流動資產總額')
        cl_k    = _find(bs_d, '流動負債合計') or _find(bs_d, '流動負債總額')
        tl_k    = _find(bs_d, '負債總額') or _find(bs_d, '負債合計')
        ta_k    = _find(bs_d, '資產總額') or _find(bs_d, '資產合計')
        eq_k    = _find(bs_d, '股東權益總額') or _find(bs_d, '權益總額')
        cash_k  = _find(bs_d, '現金') or _find(bs_d, '約當現金')
        ocf_k   = _find(cf_d, '營業活動') or _find(cf_d, '來自營運')
        capex_k = _find(cf_d, '資本支出') or _find(cf_d, '取得不動產')

        rev   = _g(is_d, rev_k,   yr)
        gp    = _g(is_d, gp_k,    yr)
        op    = _g(is_d, op_k,    yr)
        ni    = _g(is_d, ni_k,    yr)
        eps   = _g(is_d, eps_k,   yr)
        ca    = _g(bs_d, ca_k,    yr)
        cl    = _g(bs_d, cl_k,    yr)
        tl    = _g(bs_d, tl_k,    yr)
        ta    = _g(bs_d, ta_k,    yr)
        eq    = _g(bs_d, eq_k,    yr)
        cash  = _g(bs_d, cash_k,  yr)
        ocf   = _g(cf_d, ocf_k,   yr)
        capex = _g(cf_d, capex_k, yr)

        fcf = (ocf + capex) if (ocf is not None and capex is not None) else ocf

        net_margin     = safe_div(ni, rev)
        asset_turnover = safe_div(rev, ta, pct=False)
        equity_mult    = safe_div(ta, eq, pct=False)
        roe_dupont     = None
        if all(x is not None for x in [net_margin, asset_turnover, equity_mult]):
            roe_dupont = net_margin / 100 * asset_turnover * equity_mult * 100

        invested_cap = (ta - cl) if (ta is not None and cl is not None) else None
        nopat = op * 0.8 if op is not None else None

        metrics[yr] = {
            'revenue': rev, 'gross_profit': gp, 'op_income': op,
            'net_income': ni, 'eps': eps,
            'cash': cash, 'current_assets': ca, 'current_liabilities': cl,
            'total_liabilities': tl, 'total_assets': ta, 'equity': eq,
            'op_cf': ocf, 'fcf': fcf,
            'gross_margin':  safe_div(gp, rev),
            'op_margin':     safe_div(op, rev),
            'net_margin':    net_margin,
            'current_ratio': safe_div(ca, cl),
            'debt_ratio':    safe_div(tl, ta),
            'roe':           safe_div(ni, eq),
            'roa':           safe_div(ni, ta),
            'roic':          safe_div(nopat, invested_cap),
            'fcf_margin':    safe_div(fcf, rev),
            'ocf_margin':    safe_div(ocf, rev),
            'roe_dupont':    roe_dupont,
            'asset_turnover': asset_turnover,
            'equity_mult':   equity_mult,
        }

    rev_vals = [(yr, metrics[yr]['revenue']) for yr in years if metrics[yr].get('revenue')]
    if len(rev_vals) >= 2:
        cagr = calc_cagr(rev_vals[-1][1], rev_vals[0][1], int(rev_vals[0][0]) - int(rev_vals[-1][0]))
        for yr in years:
            metrics[yr]['rev_cagr'] = cagr
    return metrics

# ─── 存活評估（求職用）────────────────────────────────────
# 核心邏輯：從財報算出「這公司 2 年後還在嗎？」的風險等級。
# 評分制：累積 risk_points，>=5 高風險，>=2 中等，<2 低風險。

def trend_arrow(metrics, years, key):
    """比較最新年 vs 最舊年，漲超過 5% 給 ↑，跌超過 5% 給 ↓，否則 →"""
    vals = [(yr, metrics[yr].get(key)) for yr in years if metrics[yr].get(key) is not None]
    if len(vals) < 2:
        return '→'
    v_new, v_old = vals[0][1], vals[-1][1]
    if v_old == 0:
        return '→'
    pct = (v_new - v_old) / abs(v_old) * 100
    return '↑' if pct > 5 else ('↓' if pct < -5 else '→')

def calc_survival_score(stock_data):
    years   = stock_data['years']
    metrics = stock_data['metrics']
    signals = []
    pts     = 0

    # 1. 自由現金流：公司能不能產生現金？連續負 FCF = 燒錢警報
    fcf_vals = [metrics[yr].get('fcf') for yr in years if metrics[yr].get('fcf') is not None]
    neg_fcf  = sum(1 for v in fcf_vals if v < 0)
    if neg_fcf >= 2:
        pts += 3; signals.append(('danger', f'自由現金流連 {neg_fcf} 年為負，持續燒錢'))
    elif neg_fcf == 1:
        pts += 1; signals.append(('warn', '自由現金流近期有一年為負'))
    elif fcf_vals:
        signals.append(('ok', f'自由現金流持續為正（{len(fcf_vals)} 年）'))

    # 2. 營收趨勢：業務在成長還是萎縮？
    rev_vals = [(yr, metrics[yr].get('revenue')) for yr in years if metrics[yr].get('revenue')]
    if len(rev_vals) >= 2:
        n    = int(rev_vals[0][0]) - int(rev_vals[-1][0])
        cagr = calc_cagr(rev_vals[-1][1], rev_vals[0][1], n)
        if cagr is not None:
            if cagr <= -5:
                pts += 2; signals.append(('danger', f'營收年均衰退 {abs(cagr):.1f}%，業務萎縮'))
            elif cagr < 0:
                pts += 1; signals.append(('warn', f'營收微幅衰退，CAGR {cagr:.1f}%'))
            elif cagr >= 10:
                signals.append(('ok', f'營收強勁成長，CAGR +{cagr:.1f}%'))
            else:
                signals.append(('ok', f'營收穩定成長，CAGR +{cagr:.1f}%'))

    # 3. 獲利能力：有沒有在賺錢？
    ni_vals = [metrics[yr].get('net_income') for yr in years if metrics[yr].get('net_income') is not None]
    neg_ni  = sum(1 for v in ni_vals if v < 0)
    if neg_ni >= 2:
        pts += 3; signals.append(('danger', f'連續 {neg_ni} 年虧損'))
    elif neg_ni == 1:
        pts += 1; signals.append(('warn', '近期有虧損年份'))
    elif ni_vals:
        signals.append(('ok', '持續獲利'))

    latest_yr = years[0] if years else None
    if latest_yr:
        # 4. 負債比率：財務槓桿合理嗎？
        dr = metrics[latest_yr].get('debt_ratio')
        if dr is not None:
            if dr > 70:
                pts += 2; signals.append(('danger', f'負債比率 {dr:.0f}%，槓桿過高'))
            elif dr > 50:
                pts += 1; signals.append(('warn', f'負債比率 {dr:.0f}%，需關注'))
            else:
                signals.append(('ok', f'負債比率 {dr:.0f}%，財務結構健全'))

        # 5. 流動比率：短期能還債嗎？
        cr = metrics[latest_yr].get('current_ratio')
        if cr is not None:
            if cr < 100:
                pts += 2; signals.append(('danger', f'流動比率 {cr:.0f}%，短期償債壓力大'))
            elif cr < 150:
                pts += 1; signals.append(('warn', f'流動比率 {cr:.0f}%，短期流動性偏緊'))
            else:
                signals.append(('ok', f'流動比率 {cr:.0f}%，短期償債能力充足'))

        # 6. 現金跑道：只在 OCF 為負時才算（正常獲利公司不適用）
        cash = metrics[latest_yr].get('cash')
        ocf  = metrics[latest_yr].get('op_cf')
        if cash and ocf is not None and ocf < 0:
            runway = cash / (abs(ocf) / 12)
            if runway < 12:
                pts += 3; signals.append(('danger', f'現金跑道僅 {runway:.0f} 個月，資金壓力極大'))
            elif runway < 24:
                pts += 1; signals.append(('warn', f'現金跑道約 {runway:.0f} 個月'))

    if pts >= 5:
        return {'level': 'high',   'label': '高風險', 'color': '#e53e3e', 'pts': pts,
                'verdict': '多項財務警示，短期財務壓力明顯。入職前務必確認現金流與負債狀況。',
                'signals': signals}
    if pts >= 2:
        return {'level': 'medium', 'label': '中等風險', 'color': '#dd6b20', 'pts': pts,
                'verdict': '財務體質尚可，部分指標需關注。建議追蹤最新一期財報後再決定。',
                'signals': signals}
    return     {'level': 'low',    'label': '低風險',   'color': '#38a169', 'pts': pts,
                'verdict': '財務體質健全，近期財務危機風險低。從財務角度可放心評估。',
                'signals': signals}

# ─── 金融檢察官（財報法醫分析）───────────────────────────────
# 根據公開財報數字，找常見的財報造假、掏空、洗錢跡象。
# 方法論：改編自 Beneish M-Score + 台灣常見案例特徵。
# 注意：這是量化警示，不是法律定論。有 flag 不代表一定有問題，無 flag 不代表絕對乾淨。

def calc_forensic_score(stock_data):
    years   = stock_data['years']
    metrics = stock_data['metrics']
    is_d    = stock_data['income_statement']
    bs_d    = stock_data['balance_sheet']
    cf_d    = stock_data['cash_flow']
    flags   = []   # (severity, category, detail)
    score   = 0    # 累計可疑分數

    # ── 1. 盈餘品質（OCF/NI ratio）────────────────────────
    # 正常公司 OCF ≥ NI（折舊等非現金費用會讓 OCF 更高）
    # OCF 持續 < NI 80% = 帳面利潤可能虛增：提前認列收入、延後認列費用
    ocf_ni_pairs = [(yr, metrics[yr].get('op_cf'), metrics[yr].get('net_income'))
                    for yr in years
                    if metrics[yr].get('op_cf') is not None and metrics[yr].get('net_income')]
    low_q = [(yr, ocf/ni) for yr, ocf, ni in ocf_ni_pairs if ni > 0 and ocf/ni < 0.8]
    if len(low_q) >= 3:
        avg = sum(r for _, r in low_q) / len(low_q)
        score += 3
        flags.append(('critical', '盈餘品質異常',
            f'連續 {len(low_q)} 年 OCF/淨利 < 80%（平均 {avg:.0%}）。'
            '淨利遠超現金流是財報操縱最典型特徵，常見手法：提前認列收入、'
            '費用遞延至下一期、應收帳款虛增。'))
    elif len(low_q) >= 2:
        score += 1
        flags.append(('warning', '盈餘品質偏低',
            f'近 {len(low_q)} 年 OCF 未達淨利 80%，獲利含金量不足，需追蹤趨勢。'))
    elif ocf_ni_pairs:
        avg = sum(ocf/ni for _, ocf, ni in ocf_ni_pairs if ni > 0) / len(ocf_ni_pairs)
        if avg >= 1.0:
            flags.append(('ok', '盈餘品質', f'OCF 平均為淨利的 {avg:.0%}，現金流品質良好。'))

    # ── 2. 有盈無現（Profitable but cash-negative）────────
    # 帳面賺錢但自由現金流持續為負 = 利潤可能是帳面數字
    pnc = [yr for yr in years
           if (metrics[yr].get('net_income') or 0) > 0
           and (metrics[yr].get('fcf') or 0) < 0]
    if len(pnc) >= 3:
        score += 3
        flags.append(('critical', '有盈無現（掏空/造假警示）',
            f'連續 {len(pnc)} 年帳面獲利但自由現金流為負（{", ".join(pnc)}）。'
            '這是台灣掏空案最常見的財報特徵之一：數字好看但公司沒有真實現金進帳。'))
    elif len(pnc) >= 2:
        score += 1
        flags.append(('warning', '有盈無現',
            f'{len(pnc)} 年出現帳面盈利但 FCF 為負，需確認資本支出是否合理。'))

    # ── 3. 資產膨脹 vs 營收（虛增資產警示）───────────────
    # 資產成長速度遠超營收 → 可能用關聯交易灌水資產、或購入無效資產轉移現金
    ta_vals  = [(yr, metrics[yr].get('total_assets')) for yr in years if metrics[yr].get('total_assets')]
    rev_vals = [(yr, metrics[yr].get('revenue'))      for yr in years if metrics[yr].get('revenue')]
    if len(ta_vals) >= 2 and len(rev_vals) >= 2:
        n = int(ta_vals[0][0]) - int(ta_vals[-1][0])
        if n > 0:
            ta_cagr  = calc_cagr(ta_vals[-1][1],  ta_vals[0][1],  n)
            rev_cagr = calc_cagr(rev_vals[-1][1], rev_vals[0][1], n)
            if ta_cagr is not None and rev_cagr is not None:
                diff = ta_cagr - rev_cagr
                if diff > 20:
                    score += 3
                    flags.append(('critical', '資產異常膨脹',
                        f'資產年均成長 {ta_cagr:.1f}%，遠超營收成長 {rev_cagr:.1f}%（差 {diff:.1f}ppt）。'
                        '常見手法：用高價向關係人購入資產、虛增存貨或預付款項、'
                        '投資空殼子公司將現金轉出。'))
                elif diff > 10:
                    score += 1
                    flags.append(('warning', '資產成長偏快',
                        f'資產（{ta_cagr:.1f}%/yr）略超營收（{rev_cagr:.1f}%/yr），建議了解資產用途。'))
                else:
                    flags.append(('ok', '資產與營收成長平衡', f'資產與營收成長比例合理。'))

    # ── 4. 應收帳款異常（Channel Stuffing）────────────────
    # AR 成長速度遠超營收 → 可能塞貨給通路（月底衝業績）或虛開發票認列收入
    ar_k = _find(bs_d, '應收帳款') or _find(bs_d, '應收票據及應收帳款') or _find(bs_d, '應收')
    if ar_k:
        ar_vals = [(yr, bs_d[ar_k].get(yr)) for yr in years if bs_d[ar_k].get(yr)]
        if len(ar_vals) >= 2 and len(rev_vals) >= 2:
            n = int(ar_vals[0][0]) - int(ar_vals[-1][0])
            if n > 0:
                ar_cagr = calc_cagr(ar_vals[-1][1], ar_vals[0][1], n)
                rev_cagr2 = calc_cagr(rev_vals[-1][1], rev_vals[0][1], n)
                if ar_cagr is not None and rev_cagr2 is not None:
                    diff = ar_cagr - rev_cagr2
                    if diff > 25:
                        score += 3
                        flags.append(('critical', '應收帳款暴增（Channel Stuffing 警示）',
                            f'應收帳款年均成長 {ar_cagr:.1f}%，遠超營收成長 {rev_cagr2:.1f}%（差 {diff:.1f}ppt）。'
                            '典型操縱手法：月底大量出貨給關聯通路衝業績，但貨品最終退回；'
                            '或對空殼客戶開立假發票認列收入。'))
                    elif diff > 15:
                        score += 1
                        flags.append(('warning', '應收帳款成長偏快',
                            f'AR 成長（{ar_cagr:.1f}%）超過營收（{rev_cagr2:.1f}%），需確認客戶信用狀況。'))
                    else:
                        flags.append(('ok', '應收帳款', '應收帳款與營收成長比例正常。'))

    # ── 5. 毛利率突然跳升（費用資本化 / 成本操縱）─────────
    # 正常公司毛利率改善緩慢；單年暴增往往代表費用被錯誤資本化或成本被推遲認列
    gm_vals = [(yr, metrics[yr].get('gross_margin')) for yr in years if metrics[yr].get('gross_margin') is not None]
    if len(gm_vals) >= 3:
        jumps = [(gm_vals[i][0], gm_vals[i][1] - gm_vals[i+1][1]) for i in range(len(gm_vals)-1)]
        worst = max(jumps, key=lambda x: x[1])
        if worst[1] > 10:
            score += 2
            flags.append(('critical', f'毛利率單年暴增（{worst[0]}）',
                f'毛利率在 {worst[0]} 年單年跳升 {worst[1]:.1f}ppt，異常大幅改善。'
                '常見原因：期末把應列費用轉入資本支出（WorldCom 手法）、'
                '壓低對供應商的成本認列、或透過關聯公司交易墊高售價。'))
        elif worst[1] > 6:
            score += 1
            flags.append(('warning', f'毛利率顯著提升（{worst[0]}）',
                f'毛利率在 {worst[0]} 年提升 {worst[1]:.1f}ppt，建議了解是否有業務結構重大改變。'))

    # ── 6. 負債突然暴增（隱藏負債被迫揭露）───────────────
    tl_vals = [(yr, metrics[yr].get('total_liabilities')) for yr in years if metrics[yr].get('total_liabilities')]
    if len(tl_vals) >= 2:
        for i in range(len(tl_vals)-1):
            yr_new, v_new = tl_vals[i]
            yr_old, v_old = tl_vals[i+1]
            if v_old and v_old > 0:
                chg = (v_new - v_old) / v_old * 100
                if chg > 50:
                    score += 2
                    flags.append(('critical', f'負債單年暴增 {chg:.0f}%（{yr_old}→{yr_new}）',
                        f'總負債從 {fmt(v_old,0)} 億暴增至 {fmt(v_new,0)} 億（+{chg:.0f}%）。'
                        '可能是表外負債被迫揭露、擔保責任轉為正式負債、'
                        '或以大量借款掩護現金轉出。'))
                    break
                elif chg > 30:
                    score += 1
                    flags.append(('warning', f'負債快速增加（{yr_old}→{yr_new}）',
                        f'總負債增加 {chg:.0f}%，建議確認資金用途是否為正常業務擴張。'))
                    break

    # ── 7. 股東權益流失（掏空指標）────────────────────────
    # 累計淨利應會反映在股東權益增加；若累計利潤高但權益反而縮水，代表大量現金被轉出
    eq_vals = [(yr, metrics[yr].get('equity')) for yr in years if metrics[yr].get('equity')]
    if len(eq_vals) >= 3:
        eq_new, eq_old = eq_vals[0][1], eq_vals[-1][1]
        total_ni = sum(metrics[yr].get('net_income') or 0 for yr in years)
        if eq_old and total_ni > 0 and eq_old > 0:
            expected = eq_old + total_ni
            gap = expected - eq_new
            gap_pct = gap / eq_old * 100
            if gap_pct > 60 and gap > 0:
                score += 3
                flags.append(('critical', '股東權益異常流失（疑似掏空）',
                    f'近 {len(years)} 年累計淨利約 {fmt(total_ni,0)} 億，'
                    f'但股東權益僅從 {fmt(eq_old,0)} 億變為 {fmt(eq_new,0)} 億（應有 {fmt(expected,0)} 億），'
                    f'差距 {fmt(gap,0)} 億（{gap_pct:.0f}%）無法以股利完全解釋。'
                    '經典掏空手法：以高額股利為名分配現金後再增資圈錢、'
                    '或透過關聯方交易將資金轉至私人帳戶。'))
            elif gap_pct > 30 and gap > 0:
                score += 1
                flags.append(('warning', '股東權益成長低於預期',
                    f'累計獲利與權益增加差距 {fmt(gap,0)} 億（{gap_pct:.0f}%），'
                    '若非高股利政策，需了解資金去向。'))

    # ── 最終裁定 ──────────────────────────────────────────
    if score >= 6:
        return {'level': 'critical', 'label': '高度可疑', 'color': '#c53030', 'score': score,
                'summary': f'共發現 {score} 點財報異常訊號，多項指標同時出現是造假的強力警示。強烈建議查閱公開資訊觀測站重大訊息、近 3 年年報附註，以及是否有大股東異常減持。',
                'flags': flags}
    if score >= 3:
        return {'level': 'warning', 'label': '部分疑點', 'color': '#c05621', 'score': score,
                'summary': f'發現 {score} 點財務異常，單獨看可能有合理解釋，但組合出現需提高警惕。建議進一步查閱年報附註與董監持股變化。',
                'flags': flags}
    return     {'level': 'clean',   'label': '未見明顯異常', 'color': '#276749', 'score': score,
                'summary': '主要財報指標未出現典型造假特徵。但量化模型有盲點（如關聯方交易、董事借款等），仍建議閱讀年報附註。',
                'flags': flags}

# ─── 主要抓取流程 ──────────────────────────────────────────

def fetch_stock(stock_id, max_years=5, use_cache=True):
    conn = init_cache()
    client_key, days = get_client_key()

    print(f"  [{stock_id}] 損益表...", end=' ', flush=True)
    is_soup = fetch_report(stock_id, 'IS_YEAR', days, client_key, conn, use_cache)
    is_d, years = parse_table(is_soup, max_years)
    print("✓")

    time.sleep(0.8)
    print(f"  [{stock_id}] 資產負債表...", end=' ', flush=True)
    bs_soup = fetch_report(stock_id, 'BS_YEAR', days, client_key, conn, use_cache)
    bs_d, _ = parse_table(bs_soup, max_years)
    print("✓")

    time.sleep(0.8)
    print(f"  [{stock_id}] 現金流量表...", end=' ', flush=True)
    cf_soup = fetch_report(stock_id, 'CF_YEAR', days, client_key, conn, use_cache)
    cf_d, _ = parse_table(cf_soup, max_years)
    print("✓")

    conn.close()
    years   = years[:max_years]
    metrics = calculate_metrics(is_d, bs_d, cf_d, years)
    return {'stock_id': stock_id, 'years': years,
            'income_statement': is_d, 'balance_sheet': bs_d,
            'cash_flow': cf_d, 'metrics': metrics}

# ─── 輔助格式化 ────────────────────────────────────────────

def fmt(v, decimals=1, suffix=''):
    return '—' if v is None else f"{v:,.{decimals}f}{suffix}"

def fmt_b(v):
    return fmt(v, 1, ' 億')

def hex_to_rgba(hex_color, alpha=0.15):
    h = hex_color.lstrip('#')
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"

def js_array(lst):
    return '[' + ','.join(str(v) if v is not None else 'null' for v in lst) + ']'

# ─── HTML CSS ─────────────────────────────────────────────

CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Microsoft JhengHei','Noto Sans TC',sans-serif; background: #f0f4f8; color: #2d3748; }
.header { background: linear-gradient(135deg,#1a365d,#2b6cb0,#3182ce); color: white; padding: 20px 32px; }
.header h1 { font-size: 1.4rem; margin-bottom: 4px; }
.header .sub { font-size: 0.82rem; opacity: 0.85; }
.tabs { display: flex; background: white; border-bottom: 2px solid #e2e8f0; padding: 0 24px; overflow-x: auto; }
.tab { padding: 12px 20px; cursor: pointer; font-size: 0.9rem; font-weight: 600; color: #718096;
  border-bottom: 3px solid transparent; margin-bottom: -2px; white-space: nowrap; }
.tab.active { color: #2b6cb0; border-bottom-color: #2b6cb0; }
.tab-content { display: none; padding: 24px 32px; }
.tab-content.active { display: block; }
.section-title { font-size: 1rem; font-weight: 700; color: #2d3748; margin: 20px 0 12px;
  padding-left: 10px; border-left: 4px solid #3182ce; }
.stock-cards { display: grid; grid-template-columns: repeat(auto-fit,minmax(220px,1fr)); gap: 16px; margin-bottom: 20px; }
.stock-summary-card { background: white; border-radius: 12px; padding: 20px;
  box-shadow: 0 2px 8px rgba(0,0,0,0.07); border-top: 4px solid #3182ce; }
.stock-summary-card h2 { font-size: 1.1rem; margin-bottom: 12px; }
.metric-row { display: flex; justify-content: space-between; padding: 4px 0;
  border-bottom: 1px solid #f0f4f8; font-size: 0.84rem; }
.metric-row:last-child { border-bottom: none; }
.metric-name { color: #718096; }
.metric-val { font-weight: 600; }
.charts-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; margin-bottom: 20px; }
.chart-card { background: white; border-radius: 10px; padding: 18px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); }
.chart-title { font-size: 0.88rem; font-weight: 700; color: #4a5568; margin-bottom: 12px; }
.chart-container { position: relative; height: 220px; }
.cmp-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; background: white;
  border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.07); }
.cmp-table th { background: #2b6cb0; color: white; padding: 10px 14px; text-align: center; }
.cmp-table th:first-child { text-align: left; }
.cmp-table td { padding: 9px 14px; text-align: right; border-bottom: 1px solid #e2e8f0; }
.cmp-table td:first-child { text-align: left; font-weight: 600; color: #4a5568; background: #f7fafc; }
.cmp-table tr:last-child td { border-bottom: none; }
.cmp-table .cat-row td { background: #ebf8ff; color: #2b6cb0; font-weight: 700; font-size: 0.82rem; }
.best  { background: #f0fff4 !important; color: #276749 !important; font-weight: 700; }
.worst { background: #fff5f5 !important; color: #9b2c2c !important; }
.insight-box { background: linear-gradient(135deg,#ebf8ff,#e6fffa); border: 1px solid #bee3f8;
  border-radius: 10px; padding: 16px 20px; margin-bottom: 18px; }
.insight-box h3 { color: #2b6cb0; font-size: 0.88rem; margin-bottom: 8px; }
.insight-box li { font-size: 0.84rem; color: #4a5568; padding: 2px 0 2px 16px; position: relative; list-style: none; }
.insight-box li::before { content: '▸'; position: absolute; left: 0; color: #3182ce; }
.survival-grid { display: grid; grid-template-columns: repeat(auto-fit,minmax(300px,1fr)); gap: 20px; margin-bottom: 24px; }
.survival-card { background: white; border-radius: 12px; padding: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.07); }
.survival-header { display: flex; align-items: center; gap: 12px; margin-bottom: 10px; }
.survival-sid { font-size: 1.2rem; font-weight: 700; }
.risk-badge { padding: 4px 14px; border-radius: 99px; color: white; font-size: 0.78rem; font-weight: 700; }
.risk-meter { height: 6px; border-radius: 99px; background: #e2e8f0; margin: 0 0 14px; overflow: hidden; }
.risk-meter-fill { height: 100%; border-radius: 99px; }
.verdict-text { font-size: 0.87rem; color: #4a5568; line-height: 1.5; padding: 12px;
  background: #f7fafc; border-radius: 8px; margin-bottom: 14px; }
.signal-list { display: flex; flex-direction: column; gap: 6px; }
.signal-item { display: flex; align-items: flex-start; gap: 8px; font-size: 0.83rem; padding: 6px 8px; border-radius: 6px; }
.signal-ok     { background: #f0fff4; color: #276749; }
.signal-warn   { background: #fffbeb; color: #744210; }
.signal-danger { background: #fff5f5; color: #9b2c2c; }
@media (max-width: 768px) { .charts-grid { grid-template-columns: 1fr; } .survival-grid { grid-template-columns: 1fr; } .forensic-grid { grid-template-columns: 1fr; } }
/* 金融檢察官 tab */
.forensic-verdict { border-radius: 10px; padding: 18px 22px; margin-bottom: 20px; border-left: 6px solid; }
.forensic-verdict.critical { background:#fff5f5; border-color:#c53030; }
.forensic-verdict.warning  { background:#fffaf0; border-color:#c05621; }
.forensic-verdict.clean    { background:#f0fff4; border-color:#276749; }
.forensic-verdict-label { font-size:1rem; font-weight:800; margin-bottom:6px; }
.forensic-verdict-text  { font-size:0.86rem; line-height:1.6; }
.forensic-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(340px,1fr)); gap:20px; margin-bottom:20px; }
.forensic-card { background:white; border-radius:12px; overflow:hidden; box-shadow:0 2px 10px rgba(0,0,0,0.08); }
.forensic-card-header { padding:14px 20px; color:white; font-weight:700; font-size:1rem; display:flex; align-items:center; gap:10px; }
.forensic-findings { padding:16px 20px; display:flex; flex-direction:column; gap:10px; }
.finding { border-radius:8px; padding:12px 14px; }
.finding.critical { background:#fff5f5; border-left:4px solid #c53030; }
.finding.warning  { background:#fffaf0; border-left:4px solid #c05621; }
.finding.ok       { background:#f0fff4; border-left:4px solid #276749; }
.finding-category { font-size:0.78rem; font-weight:800; text-transform:uppercase; letter-spacing:0.05em; margin-bottom:5px; }
.finding.critical .finding-category { color:#c53030; }
.finding.warning  .finding-category { color:#c05621; }
.finding.ok       .finding-category { color:#276749; }
.finding-detail { font-size:0.82rem; line-height:1.6; color:#4a5568; }
.finding.critical .finding-detail { color:#742a2a; }
"""

# ─── HTML 生成 ─────────────────────────────────────────────

def build_comparison_html(all_stocks):
    stock_ids = list(all_stocks.keys())
    n         = len(stock_ids)
    all_years = all_stocks[stock_ids[0]]['years']
    latest_yr = all_years[0] if all_years else '—'
    color_map = {sid: STOCK_COLORS[i % len(STOCK_COLORS)] for i, sid in enumerate(stock_ids)}

    def m(sid, yr, key):
        return all_stocks[sid]['metrics'].get(yr, {}).get(key)

    # Best/Worst 著色：數值相同時不標色，避免誤導
    def best_worst_tds(key, yr, higher_better=True, fmt_fn=None):
        vals  = {sid: m(sid, yr, key) for sid in stock_ids}
        valid = {sid: v for sid, v in vals.items() if v is not None}
        best_v = (max if higher_better else min)(valid.values(), default=None) if valid else None
        worst_v= (min if higher_better else max)(valid.values(), default=None) if valid else None
        all_same = (best_v == worst_v)
        tds = []
        for sid in stock_ids:
            v = vals[sid]
            d = fmt_fn(v) if (fmt_fn and v is not None) else (fmt(v) if v is not None else '—')
            cls = ''
            if v is not None and not all_same:
                if v == best_v:  cls = ' class="best"'
                elif v == worst_v: cls = ' class="worst"'
            tds.append(f'<td{cls}>{d}</td>')
        return ''.join(tds)

    # Summary cards
    cards_html = ''
    for sid in stock_ids:
        color = color_map[sid]
        yr    = latest_yr
        rows  = ''.join(
            f'<div class="metric-row"><span class="metric-name">{name}</span><span class="metric-val">{val}</span></div>'
            for name, val in [
                ('營業收入', fmt_b(m(sid, yr, 'revenue'))),
                ('毛利率',   fmt(m(sid, yr, 'gross_margin'), suffix='%')),
                ('稅後淨利率', fmt(m(sid, yr, 'net_margin'), suffix='%')),
                ('EPS (元)', fmt(m(sid, yr, 'eps'), 2)),
                ('ROE',      fmt(m(sid, yr, 'roe'), suffix='%')),
                ('負債比率', fmt(m(sid, yr, 'debt_ratio'), suffix='%')),
            ]
        )
        cards_html += f'<div class="stock-summary-card" style="border-top-color:{color}"><h2 style="color:{color}">🏢 {sid}</h2>{rows}</div>'

    # Chart dataset builders — 修正：rev 用 hex_to_rgba，不再有 Python ternary
    rev_ds = ','.join(
        f'{{"label":"{sid} 營收","data":{js_array([m(sid,yr,"revenue") for yr in all_years])},'
        f'"backgroundColor":"{hex_to_rgba(color_map[sid])}","borderColor":"{color_map[sid]}","borderWidth":2}}'
        for sid in stock_ids
    )
    fcf_ds = ','.join(
        f'{{"label":"{sid} FCF","data":{js_array([m(sid,yr,"fcf") for yr in all_years])},'
        f'"backgroundColor":"{hex_to_rgba(color_map[sid],0.7)}","borderColor":"{color_map[sid]}","borderWidth":2}}'
        for sid in stock_ids
    )
    def line_ds(key, suffix):
        return ','.join(
            f'{{"label":"{sid} {suffix}","data":{js_array([m(sid,yr,key) for yr in all_years])},'
            f'"borderColor":"{color_map[sid]}","backgroundColor":"{color_map[sid]}","pointRadius":5,"tension":0.3,"fill":false}}'
            for sid in stock_ids
        )

    # Comparison table
    headers = ''.join(f'<th style="color:{color_map[sid]};background:#1a365d;">{sid}</th>' for sid in stock_ids)
    def all_yr_rows(label, key, higher=True, fmt_fn=None):
        body = f'<tr class="cat-row"><td colspan="{n+1}">📅 {label}</td></tr>'
        for yr in all_years:
            body += f'<tr><td style="padding-left:24px;color:#718096;">{yr}</td>{best_worst_tds(key,yr,higher,fmt_fn)}</tr>'
        return body

    cmp_html = f"""<table class="cmp-table">
<thead><tr><th>指標 / 年度</th>{headers}</tr></thead><tbody>
{all_yr_rows("營業收入 (億元)","revenue",True,lambda v:fmt(v,1))}
{all_yr_rows("毛利率 (%)","gross_margin",True,lambda v:fmt(v,1,'%'))}
{all_yr_rows("營業利益率 (%)","op_margin",True,lambda v:fmt(v,1,'%'))}
{all_yr_rows("稅後淨利率 (%)","net_margin",True,lambda v:fmt(v,1,'%'))}
{all_yr_rows("EPS (元)","eps",True,lambda v:fmt(v,2))}
{all_yr_rows("ROE (%)","roe",True,lambda v:fmt(v,1,'%'))}
{all_yr_rows("ROIC (%)","roic",True,lambda v:fmt(v,1,'%'))}
{all_yr_rows("流動比率 (%)","current_ratio",True,lambda v:fmt(v,0,'%'))}
{all_yr_rows("負債比率 (%)","debt_ratio",False,lambda v:fmt(v,1,'%'))}
{all_yr_rows("自由現金流 (億元)","fcf",True,lambda v:fmt(v,1))}
</tbody></table>"""

    # DuPont
    dupont_rows = ''
    for sid in stock_ids:
        yr = latest_yr
        at = (fmt(m(sid,yr,'asset_turnover'),3)+'x') if m(sid,yr,'asset_turnover') else '—'
        em = (fmt(m(sid,yr,'equity_mult'),2)+'x') if m(sid,yr,'equity_mult') else '—'
        dupont_rows += (f'<tr><td style="color:{color_map[sid]};font-weight:700;">{sid}</td>'
                        f'<td>{fmt(m(sid,yr,"net_margin"),1,"%")}</td><td>{at}</td><td>{em}</td>'
                        f'<td>{fmt(m(sid,yr,"roe"),1,"%")}</td><td>{fmt(m(sid,yr,"rev_cagr"),1,"%")}</td></tr>')

    # Survival tab
    ICONS = {'ok': '✅', 'warn': '⚠️', 'danger': '🔴'}
    survival_cards = ''
    for sid in stock_ids:
        score  = calc_survival_score(all_stocks[sid])
        color  = color_map[sid]
        meter  = min(100, score['pts'] / 8 * 100)
        sigs   = ''.join(
            f'<div class="signal-item signal-{lvl}"><span>{ICONS[lvl]}</span><span>{msg}</span></div>'
            for lvl, msg in score['signals']
        )
        survival_cards += f"""<div class="survival-card">
  <div class="survival-header">
    <span class="survival-sid" style="color:{color}">{sid}</span>
    <span class="risk-badge" style="background:{score['color']}">{score['label']}</span>
  </div>
  <div class="risk-meter"><div class="risk-meter-fill" style="width:{meter:.0f}%;background:{score['color']}"></div></div>
  <div class="verdict-text">{score['verdict']}</div>
  <div class="signal-list">{sigs}</div>
</div>"""

    # Forensic tab — 金融檢察官
    F_ICONS = {'critical': '🔴', 'warning': '⚠️', 'ok': '✅'}
    forensic_verdicts = ''
    forensic_cards    = ''
    for sid in stock_ids:
        fscore = calc_forensic_score(all_stocks[sid])
        color  = color_map[sid]
        forensic_verdicts += f"""<div class="forensic-verdict {fscore['level']}">
  <div class="forensic-verdict-label" style="color:{fscore['color']}">
    ⚖️ {sid} — 裁定：{fscore['label']}（可疑分 {fscore['score']} 點）
  </div>
  <div class="forensic-verdict-text">{fscore['summary']}</div>
</div>"""
        findings_html = ''.join(
            f'<div class="finding {sev}"><div class="finding-category">{F_ICONS[sev]} {cat}</div>'
            f'<div class="finding-detail">{detail}</div></div>'
            for sev, cat, detail in fscore['flags']
        )
        forensic_cards += f"""<div class="forensic-card">
  <div class="forensic-card-header" style="background:{color}">
    🔍 {sid} 財報偵查報告
  </div>
  <div class="forensic-findings">{findings_html}</div>
</div>"""

    title   = ' vs '.join(stock_ids)
    fetched = date.today().isoformat()
    labels  = str(all_years).replace("'", '"')

    return f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>財務比較: {title}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.0/dist/chart.umd.min.js"></script>
<style>{CSS}</style>
</head>
<body>
<div class="header">
  <h1>📊 台灣股票財務比較分析</h1>
  <div class="sub">分析標的：{title}｜資料來源：Goodinfo.tw｜分析日期：{fetched}｜金額單位：億元</div>
</div>
<div class="tabs">
  <div class="tab active"   onclick="switchTab('overview',this)">🏠 概覽</div>
  <div class="tab" onclick="switchTab('ops',this)">📈 營運</div>
  <div class="tab" onclick="switchTab('profit',this)">💰 獲利</div>
  <div class="tab" onclick="switchTab('finance',this)">🏦 財務健全度</div>
  <div class="tab" onclick="switchTab('dupont',this)">🔬 進階分析</div>
  <div class="tab" onclick="switchTab('survival',this)">🛡️ 存活評估</div>
  <div class="tab" onclick="switchTab('forensic',this)">⚖️ 金融檢察官</div>
</div>

<div id="overview" class="tab-content active">
  <div class="section-title">各股 {latest_yr} 年度快速摘要</div>
  <div class="stock-cards">{cards_html}</div>
  <div class="section-title">全期間完整比較表</div>
  {cmp_html}
  <p style="font-size:0.75rem;color:#718096;margin-top:8px;">🟢 綠底=同期最佳　🔴 紅底=同期最差（數值相同時不標色）</p>
</div>

<div id="ops" class="tab-content">
  <div class="charts-grid">
    <div class="chart-card"><div class="chart-title">營業收入趨勢 (億元)</div><div class="chart-container"><canvas id="revChart"></canvas></div></div>
    <div class="chart-card"><div class="chart-title">毛利率 (%)</div><div class="chart-container"><canvas id="gmChart"></canvas></div></div>
    <div class="chart-card"><div class="chart-title">營業利益率 (%)</div><div class="chart-container"><canvas id="opChart"></canvas></div></div>
    <div class="chart-card"><div class="chart-title">稅後淨利率 (%)</div><div class="chart-container"><canvas id="nmChart"></canvas></div></div>
  </div>
</div>
<div id="profit" class="tab-content">
  <div class="charts-grid">
    <div class="chart-card"><div class="chart-title">ROE 趨勢 (%)</div><div class="chart-container"><canvas id="roeChart"></canvas></div></div>
    <div class="chart-card"><div class="chart-title">自由現金流 (億元)</div><div class="chart-container"><canvas id="fcfChart"></canvas></div></div>
  </div>
</div>
<div id="finance" class="tab-content">
  <div class="charts-grid">
    <div class="chart-card"><div class="chart-title">負債比率 (%)</div><div class="chart-container"><canvas id="drChart"></canvas></div></div>
  </div>
</div>
<div id="dupont" class="tab-content">
  <div class="section-title">DuPont 三因子拆解（{latest_yr}）</div>
  <div class="insight-box"><h3>📖 DuPont 公式說明</h3><ul>
    <li>ROE = 淨利率 × 資產周轉率 × 財務槓桿倍數</li>
    <li>淨利率高 → 獲利能力強（品牌/技術護城河）</li>
    <li>資產周轉率高 → 資產使用效率高（零售/製造業優勢）</li>
    <li>財務槓桿高 → 借款放大獲利（金融業常見，注意風險）</li>
  </ul></div>
  <table class="cmp-table">
    <thead><tr><th>股票</th><th>淨利率</th><th>資產周轉率</th><th>財務槓桿</th><th>ROE（實際）</th><th>營收CAGR</th></tr></thead>
    <tbody>{dupont_rows}</tbody>
  </table>
</div>
<div id="survival" class="tab-content">
  <div class="insight-box"><h3>🛡️ 這頁回答什麼問題？</h3><ul>
    <li>這間公司 2 年後還在嗎？財務危機風險高不高？</li>
    <li>現金流健康嗎？短期能不能還債？</li>
    <li>業務是在成長還是萎縮？</li>
  </ul></div>
  <div class="survival-grid">{survival_cards}</div>
  <p style="font-size:0.75rem;color:#718096;">⚠️ 本評估基於公開財報，不構成投資或求職建議。建議搭配最新季報及面試資訊綜合判斷。</p>
</div>
<div id="forensic" class="tab-content">
  <div class="insight-box"><h3>⚖️ 金融檢察官在找什麼？</h3><ul>
    <li>盈餘品質：淨利 vs 現金流是否吻合？（不吻合 = 可能帳面造假）</li>
    <li>有盈無現：帳面賺錢但持續無現金 = 台灣掏空案最典型特徵</li>
    <li>資產膨脹：資產成長遠超營收 = 可能用關聯交易虛增資產</li>
    <li>應收帳款暴增：塞貨給通路或虛開發票（Channel Stuffing）</li>
    <li>毛利率突跳：費用資本化、成本操縱</li>
    <li>股東權益流失：利潤無法解釋的消失 = 疑似掏空</li>
  </ul></div>
  <div class="section-title">各股裁定摘要</div>
  {forensic_verdicts}
  <div class="section-title">詳細偵查報告</div>
  <div class="forensic-grid">{forensic_cards}</div>
  <p style="font-size:0.75rem;color:#718096;margin-top:12px;">
    ⚠️ 本分析為量化財報異常偵測，改編自 Beneish M-Score 方法論。有 flag 不等於定罪，無 flag 不等於清白。
    確認可疑訊號請查閱：<strong>公開資訊觀測站 → 重大訊息 / 財務報告 / 關係人交易</strong>。
  </p>
</div>

<script>
function switchTab(name,el){{
  document.querySelectorAll('.tab-content').forEach(e=>e.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(e=>e.classList.remove('active'));
  document.getElementById(name).classList.add('active');
  el.classList.add('active');
}}
const labels={labels};
const lineOpts={{responsive:true,maintainAspectRatio:false,
  plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}},
  scales:{{x:{{grid:{{display:false}}}},y:{{ticks:{{callback:v=>v+'%'}}}}}}}};
const barOpts={{responsive:true,maintainAspectRatio:false,
  plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}},
  scales:{{x:{{grid:{{display:false}}}},y:{{grid:{{color:'rgba(0,0,0,0.05)'}}}}}}}};
new Chart(document.getElementById('revChart'),{{type:'bar',data:{{labels,datasets:[{rev_ds}]}},options:barOpts}});
new Chart(document.getElementById('gmChart'), {{type:'line',data:{{labels,datasets:[{line_ds("gross_margin","毛利率")}]}},options:lineOpts}});
new Chart(document.getElementById('opChart'), {{type:'line',data:{{labels,datasets:[{line_ds("op_margin","營業利益率")}]}},options:lineOpts}});
new Chart(document.getElementById('nmChart'), {{type:'line',data:{{labels,datasets:[{line_ds("net_margin","淨利率")}]}},options:lineOpts}});
new Chart(document.getElementById('roeChart'),{{type:'line',data:{{labels,datasets:[{line_ds("roe","ROE")}]}},options:lineOpts}});
new Chart(document.getElementById('fcfChart'),{{type:'bar',data:{{labels,datasets:[{fcf_ds}]}},options:barOpts}});
new Chart(document.getElementById('drChart'), {{type:'line',data:{{labels,datasets:[{line_ds("debt_ratio","負債比率")}]}},options:lineOpts}});
</script>
</body></html>"""

# ─── Excel 生成 ────────────────────────────────────────────

def build_excel(all_stocks):
    if not HAS_EXCEL:
        print("⚠️  openpyxl 未安裝，跳過 Excel。執行: pip install openpyxl")
        return None

    wb       = openpyxl.Workbook()
    sids     = list(all_stocks.keys())
    latest_yr= all_stocks[sids[0]]['years'][0]

    # ── 共用 styles ──
    hdr_fill  = PatternFill("solid", fgColor="1A365D")
    cat_fill  = PatternFill("solid", fgColor="EBF8FF")
    ok_fill   = PatternFill("solid", fgColor="F0FFF4")
    warn_fill = PatternFill("solid", fgColor="FFFBEB")
    danger_fill=PatternFill("solid", fgColor="FFF5F5")
    best_fill = PatternFill("solid", fgColor="C6F6D5")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    bold      = Font(bold=True)
    right     = Alignment(horizontal='right')
    center    = Alignment(horizontal='center', wrap_text=True)

    def hdr(ws, row, col, val):
        c = ws.cell(row, col, val)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

    SIGNAL_ICON = {'ok': '✅', 'warn': '⚠️', 'danger': '🔴'}
    ARROW_COLORS= {'↑': '276749', '↓': '9B2C2C', '→': '718096'}

    # ══ Sheet 1: 求職評估 ══════════════════════════════════
    ws = wb.active
    ws.title = "求職評估"
    ws.cell(1,1,"求職存活評估 — 這間公司 2 年後還在嗎？").font = Font(bold=True, size=13)
    ws.merge_cells(f'A1:{get_column_letter(len(sids)+1)}1')

    hdr(ws, 2, 1, "評估項目")
    for i, sid in enumerate(sids, 2):
        hdr(ws, 2, i, sid)

    scores = {sid: calc_survival_score(all_stocks[sid]) for sid in sids}
    RISK_COLORS = {'low': 'C6F6D5', 'medium': 'FEEBC8', 'high': 'FED7D7'}
    RISK_FONTS  = {'low': '276749', 'medium': '744210', 'high': '9B2C2C'}

    # Risk level row
    ws.cell(3, 1, "存活風險等級").font = bold
    for i, sid in enumerate(sids, 2):
        sc = scores[sid]
        c  = ws.cell(3, i, sc['label'])
        c.fill      = PatternFill("solid", fgColor=RISK_COLORS[sc['level']])
        c.font      = Font(bold=True, color=RISK_FONTS[sc['level']])
        c.alignment = center

    # Verdict row
    ws.cell(4, 1, "評估結論").font = bold
    ws.row_dimensions[4].height = 40
    for i, sid in enumerate(sids, 2):
        c = ws.cell(4, i, scores[sid]['verdict'])
        c.alignment = center

    # Signals
    ws.cell(6, 1, "── 財務信號 ──").font = Font(bold=True, color="2B6CB0")
    ws.cell(6, 1).fill = cat_fill
    ws.merge_cells(f'A6:{get_column_letter(len(sids)+1)}6')

    max_sigs = max(len(scores[sid]['signals']) for sid in sids)
    for row_i in range(max_sigs):
        r = 7 + row_i
        ws.row_dimensions[r].height = 18
        for i, sid in enumerate(sids, 2):
            sigs = scores[sid]['signals']
            if row_i < len(sigs):
                lvl, msg = sigs[row_i]
                c = ws.cell(r, i, f"{SIGNAL_ICON[lvl]} {msg}")
                c.fill = ok_fill if lvl=='ok' else (warn_fill if lvl=='warn' else danger_fill)
                c.font = Font(color=('276749' if lvl=='ok' else ('744210' if lvl=='warn' else '9B2C2C')), size=9)
                c.alignment = Alignment(wrap_text=True)

    # Key metric trends
    row = 7 + max_sigs + 1
    ws.cell(row, 1, "── 關鍵指標趨勢 ──").font = Font(bold=True, color="2B6CB0")
    ws.cell(row, 1).fill = cat_fill
    ws.merge_cells(f'A{row}:{get_column_letter(len(sids)+1)}{row}')
    row += 1

    trend_metrics = [
        ("營收 CAGR (%)", "rev_cagr", True),
        ("毛利率 (%)", "gross_margin", True),
        ("稅後淨利率 (%)", "net_margin", True),
        ("ROE (%)", "roe", True),
        ("自由現金流 (億元)", "fcf", True),
        ("負債比率 (%)", "debt_ratio", False),
        ("流動比率 (%)", "current_ratio", True),
    ]
    for label, key, higher_better in trend_metrics:
        ws.cell(row, 1, label).font = bold
        for i, sid in enumerate(sids, 2):
            years   = all_stocks[sid]['years']
            metrics = all_stocks[sid]['metrics']
            v       = metrics.get(latest_yr, {}).get(key)
            arrow   = trend_arrow(metrics, years, key)
            display = (fmt(v, 1) + ' ' + arrow) if v is not None else '—'
            c = ws.cell(row, i, display)
            c.alignment = right
            if v is not None:
                c.font = Font(color=ARROW_COLORS[arrow])
        row += 1

    ws.column_dimensions['A'].width = 22
    for i in range(2, len(sids) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 26

    # ══ Sheet 2: 比較摘要 ══════════════════════════════════
    ws2 = wb.create_sheet("比較摘要")
    ws2.cell(1, 1, f"財務比較摘要 — {' vs '.join(sids)} — {latest_yr}年度").font = Font(bold=True, size=13)
    ws2.merge_cells(f'A1:{get_column_letter(len(sids)+1)}1')

    hdr(ws2, 2, 1, "指標（含趨勢箭頭）")
    for i, sid in enumerate(sids, 2):
        hdr(ws2, 2, i, sid)

    metrics_rows = [
        ("─ 規模 ─", None, None, None),
        ("營業收入 (億元)", "revenue", True, lambda v: fmt(v,1)),
        ("稅後淨利 (億元)", "net_income", True, lambda v: fmt(v,1)),
        ("自由現金流 (億元)", "fcf", True, lambda v: fmt(v,1)),
        ("─ 獲利能力 ─", None, None, None),
        ("毛利率 (%)", "gross_margin", True, lambda v: fmt(v,1)),
        ("營業利益率 (%)", "op_margin", True, lambda v: fmt(v,1)),
        ("稅後淨利率 (%)", "net_margin", True, lambda v: fmt(v,1)),
        ("EPS (元)", "eps", True, lambda v: fmt(v,2)),
        ("─ 報酬率 ─", None, None, None),
        ("ROE (%)", "roe", True, lambda v: fmt(v,1)),
        ("ROA (%)", "roa", True, lambda v: fmt(v,1)),
        ("ROIC (%)", "roic", True, lambda v: fmt(v,1)),
        ("─ 財務健全 ─", None, None, None),
        ("流動比率 (%)", "current_ratio", True, lambda v: fmt(v,0)),
        ("負債比率 (%)", "debt_ratio", False, lambda v: fmt(v,1)),
        ("─ 成長 ─", None, None, None),
        ("營收 CAGR (%)", "rev_cagr", True, lambda v: fmt(v,1)),
    ]

    r = 3
    for label, key, higher_better, fmt_fn in metrics_rows:
        if key is None:
            c = ws2.cell(r, 1, label)
            c.font = Font(bold=True, color="2B6CB0"); c.fill = cat_fill
            ws2.merge_cells(f'A{r}:{get_column_letter(len(sids)+1)}{r}')
            r += 1; continue

        ws2.cell(r, 1, label).font = bold
        vals = {sid: all_stocks[sid]['metrics'].get(latest_yr, {}).get(key) for sid in sids}
        valid = {sid: v for sid, v in vals.items() if v is not None}
        best_v = (max if higher_better else min)(valid.values(), default=None) if valid else None

        for i, sid in enumerate(sids, 2):
            years  = all_stocks[sid]['years']
            mets   = all_stocks[sid]['metrics']
            v      = vals[sid]
            arrow  = trend_arrow(mets, years, key)
            display= (fmt_fn(v) + ' ' + arrow) if (fmt_fn and v is not None) else '—'
            c = ws2.cell(r, i, display)
            c.alignment = right
            c.font = Font(color=ARROW_COLORS.get(arrow, '2d3748'))
            if v is not None and v == best_v:
                c.fill = best_fill
                c.font = Font(bold=True, color='276749')
        r += 1

    ws2.column_dimensions['A'].width = 22
    for i in range(2, len(sids) + 2):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    # ══ Per-stock detail sheets ════════════════════════════
    for sid in sids:
        data  = all_stocks[sid]
        years = data['years']
        ws3   = wb.create_sheet(f"{sid}詳細")

        ws3.cell(1, 1, f"{sid} 詳細財務數據").font = Font(bold=True, size=12)
        ws3.merge_cells(f'A1:{get_column_letter(len(years)+1)}1')
        hdr(ws3, 2, 1, "項目")
        for i, yr in enumerate(years, 2):
            hdr(ws3, 2, i, yr)

        r = 3
        for sec_name, table in [("損益表", data['income_statement']),
                                  ("資產負債表", data['balance_sheet']),
                                  ("現金流量表", data['cash_flow'])]:
            c = ws3.cell(r, 1, f"── {sec_name} ──")
            c.font = Font(bold=True, color="2B6CB0"); c.fill = cat_fill
            ws3.merge_cells(f'A{r}:{get_column_letter(len(years)+1)}{r}')
            r += 1
            for field, yr_vals in list(table.items())[:40]:
                ws3.cell(r, 1, field).font = bold
                for i, yr in enumerate(years, 2):
                    v = yr_vals.get(yr)
                    if v is not None:
                        c = ws3.cell(r, i, v)
                        c.alignment = right; c.number_format = '#,##0.00'
                r += 1

        c = ws3.cell(r, 1, "── 計算指標 ──")
        c.font = Font(bold=True, color="2B6CB0"); c.fill = cat_fill
        ws3.merge_cells(f'A{r}:{get_column_letter(len(years)+1)}{r}')
        r += 1
        for label, key in [("毛利率 (%)","gross_margin"),("營業利益率 (%)","op_margin"),
                            ("淨利率 (%)","net_margin"),("EPS (元)","eps"),
                            ("ROE (%)","roe"),("ROA (%)","roa"),("ROIC (%)","roic"),
                            ("流動比率 (%)","current_ratio"),("負債比率 (%)","debt_ratio"),
                            ("自由現金流 (億)","fcf"),("FCF Margin (%)","fcf_margin")]:
            ws3.cell(r, 1, label).font = bold
            for i, yr in enumerate(years, 2):
                v = data['metrics'].get(yr, {}).get(key)
                if v is not None:
                    c = ws3.cell(r, i, round(v,2))
                    c.alignment = right; c.number_format = '#,##0.00'
            r += 1

        ws3.column_dimensions['A'].width = 24
        for i in range(2, len(years)+2):
            ws3.column_dimensions[get_column_letter(i)].width = 12

    # ══ Sheet 4: 金融檢察官 ════════════════════════════════
    ws4 = wb.create_sheet("金融檢察官")
    ws4.cell(1,1,"⚖️ 金融檢察官 — 財報異常偵測").font = Font(bold=True, size=13, color="C53030")
    ws4.merge_cells(f'A1:{get_column_letter(len(sids)+1)}1')

    hdr(ws4, 2, 1, "偵查項目")
    for i, sid in enumerate(sids, 2):
        hdr(ws4, 2, i, sid)

    fscores  = {sid: calc_forensic_score(all_stocks[sid]) for sid in sids}
    F_FILLS  = {'critical': PatternFill("solid", fgColor="FED7D7"),
                'warning':  PatternFill("solid", fgColor="FEEBC8"),
                'clean':    PatternFill("solid", fgColor="C6F6D5")}
    F_FONTS  = {'critical': '9B2C2C', 'warning': '744210', 'clean': '276749'}

    # 裁定列
    ws4.cell(3, 1, "整體裁定").font = bold
    for i, sid in enumerate(sids, 2):
        fs = fscores[sid]
        c  = ws4.cell(3, i, f"{fs['label']}（{fs['score']} 點）")
        c.fill = F_FILLS[fs['level']]; c.font = Font(bold=True, color=F_FONTS[fs['level']]); c.alignment = center

    # 裁定摘要
    ws4.row_dimensions[4].height = 60
    ws4.cell(4, 1, "裁定摘要").font = bold
    for i, sid in enumerate(sids, 2):
        c = ws4.cell(4, i, fscores[sid]['summary'])
        c.alignment = Alignment(wrap_text=True, vertical='top')

    # 各項偵查結果
    r = 6
    c_hdr = ws4.cell(r, 1, "── 詳細偵查項目 ──")
    c_hdr.font = Font(bold=True, color="C53030"); c_hdr.fill = PatternFill("solid", fgColor="FED7D7")
    ws4.merge_cells(f'A{r}:{get_column_letter(len(sids)+1)}{r}')
    r += 1

    # Collect all unique categories across all stocks
    all_cats = []
    seen = set()
    for sid in sids:
        for sev, cat, detail in fscores[sid]['flags']:
            if cat not in seen:
                all_cats.append(cat)
                seen.add(cat)

    for cat in all_cats:
        ws4.cell(r, 1, cat).font = Font(bold=True, size=9)
        ws4.row_dimensions[r].height = 45
        for i, sid in enumerate(sids, 2):
            flag = next(((s,d) for s,c,d in fscores[sid]['flags'] if c==cat), None)
            if flag:
                sev, detail = flag
                icon = {'critical':'🔴','warning':'⚠️','ok':'✅'}[sev]
                c = ws4.cell(r, i, f"{icon} {detail}")
                c.fill = F_FILLS.get(sev, PatternFill("solid", fgColor="F7FAFC"))
                c.font = Font(color=F_FONTS.get(sev, '2d3748'), size=9)
                c.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1

    ws4.cell(r+1, 1, "⚠️ 量化偵測有盲點，請以公開資訊觀測站重大訊息、關係人交易揭露、董監持股為輔助查證依據。").font = Font(italic=True, color="718096", size=9)
    ws4.merge_cells(f'A{r+1}:{get_column_letter(len(sids)+1)}{r+1}')

    ws4.column_dimensions['A'].width = 24
    for i in range(2, len(sids)+2):
        ws4.column_dimensions[get_column_letter(i)].width = 40

    out = OUTPUT_DIR / f"{'_vs_'.join(sids)}_analysis.xlsx"
    wb.save(out)
    return out

# ─── CLI 入口 ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='台灣股票財務分析工具')
    parser.add_argument('stocks', nargs='+', help='股票代碼 (1-5支)')
    parser.add_argument('--years', type=int, default=5, help='分析年數 (預設5年)')
    parser.add_argument('--no-cache', action='store_true', help='強制重新抓取')
    parser.add_argument('--no-excel', action='store_true', help='不產生 Excel')
    args = parser.parse_args()

    stock_ids = args.stocks[:5]
    use_cache = not args.no_cache

    print(f"\n🚀 台灣股票財務分析 — {' / '.join(stock_ids)}")
    print(f"   分析年數: {args.years} 年  |  快取: {'開啟' if use_cache else '關閉'}\n")

    all_stocks = {}
    for sid in stock_ids:
        print(f"📥 抓取 {sid} 資料...")
        try:
            all_stocks[sid] = fetch_stock(sid, max_years=args.years, use_cache=use_cache)
        except Exception as e:
            print(f"  ❌ {sid} 失敗: {e}")

    if not all_stocks:
        print("❌ 無資料可分析"); sys.exit(1)

    tag   = '_vs_'.join(all_stocks.keys())
    today = date.today().isoformat()

    print(f"\n📊 生成 HTML 報告...")
    html      = build_comparison_html(all_stocks)
    html_path = OUTPUT_DIR / f"{tag}_{today}.html"
    html_path.write_text(html, encoding='utf-8')
    print(f"   ✅ {html_path}")

    if not args.no_excel:
        print(f"📋 生成 Excel 報表...")
        xlsx_path = build_excel(all_stocks)
        if xlsx_path:
            print(f"   ✅ {xlsx_path}")

    print(f"\n✨ 完成！檔案在 output/ 目錄")

if __name__ == '__main__':
    main()
